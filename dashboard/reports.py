from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def build_statistics_excel(filters, statistics):
    output = BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Dashboard'

    title_fill = PatternFill('solid', fgColor='1A6B9A')
    title_font = Font(color='FFFFFF', bold=True)

    sheet['A1'] = 'Reporte de estadisticas del dashboard'
    sheet['A1'].font = Font(size=14, bold=True)
    sheet['A3'] = 'Periodo'
    sheet['B3'] = filters['periodo_label']
    sheet['A4'] = 'Desde'
    sheet['B4'] = filters['fecha_inicio'].isoformat()
    sheet['A5'] = 'Hasta'
    sheet['B5'] = filters['fecha_fin'].isoformat()

    sheet.append([])
    sheet.append(['Indicador', 'Valor'])
    for cell in sheet[7]:
        cell.fill = title_fill
        cell.font = title_font

    labels = _summary_labels()
    for key, label in labels:
        sheet.append([label, statistics['resumen'][key]])

    sheet.append([])
    sheet.append(['Citas por especialidad', 'Total'])
    for cell in sheet[sheet.max_row]:
        cell.fill = title_fill
        cell.font = title_font

    chart = statistics['charts']['citas_especialidad']
    for label, value in zip(chart['labels'], chart['data']):
        sheet.append([label, value])

    for column_cells in sheet.columns:
        sheet.column_dimensions[column_cells[0].column_letter].width = 28

    workbook.save(output)
    output.seek(0)
    return output


def build_statistics_pdf(filters, statistics):
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph('Reporte de estadisticas del dashboard', styles['Title']),
        Paragraph(
            f"Periodo: {filters['periodo_label']} "
            f"({filters['fecha_inicio'].isoformat()} a {filters['fecha_fin'].isoformat()})",
            styles['Normal'],
        ),
        Spacer(1, 16),
    ]

    summary_data = [['Indicador', 'Valor']]
    for key, label in _summary_labels():
        summary_data.append([label, str(statistics['resumen'][key])])
    elements.append(_styled_table(summary_data))
    elements.append(Spacer(1, 16))

    chart = statistics['charts']['citas_especialidad']
    specialty_data = [['Especialidad', 'Citas']]
    specialty_data.extend([label, str(value)] for label, value in zip(chart['labels'], chart['data']))
    elements.append(Paragraph('Citas por especialidad', styles['Heading2']))
    elements.append(_styled_table(specialty_data))

    doc.build(elements)
    output.seek(0)
    return output


def _summary_labels():
    return [
        ('total_pacientes', 'Total de pacientes'),
        ('pacientes_activos', 'Pacientes activos'),
        ('pacientes_inactivos', 'Pacientes inactivos'),
        ('pacientes_nuevos', 'Pacientes nuevos en periodo'),
        ('total_medicos', 'Total de medicos'),
        ('medicos_activos', 'Medicos activos'),
        ('total_citas', 'Citas en periodo'),
        ('citas_hoy', 'Citas del dia'),
        ('citas_pendientes', 'Citas pendientes'),
        ('citas_confirmadas', 'Citas confirmadas'),
        ('citas_completadas', 'Citas completadas'),
        ('citas_canceladas', 'Citas canceladas'),
    ]


def _styled_table(data):
    table = Table(data, hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A6B9A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D9E2EC')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    return table
