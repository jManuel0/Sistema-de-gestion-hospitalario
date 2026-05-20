"""
Utilidades para exportación de datos:
- Excel con openpyxl
- PDF con reportlab
"""

import io
from datetime import date

# ─── EXCEL ───────────────────────────────────────────────────────────────────

def exportar_pacientes_excel(pacientes):
    """
    Genera un archivo Excel con la lista de pacientes.
    Retorna un objeto BytesIO listo para enviar como respuesta HTTP.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Instala openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pacientes"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a6b9a", end_color="1a6b9a", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:J1')
    titulo_cell = ws['A1']
    titulo_cell.value = f"Lista de Pacientes - Generado el {date.today().strftime('%d/%m/%Y')}"
    titulo_cell.font = Font(bold=True, size=13, color="1a6b9a")
    titulo_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 25

    # Encabezados
    encabezados = [
        'DNI', 'Nombre', 'Apellido', 'Fecha Nacimiento',
        'Edad', 'Género', 'Tipo Sangre', 'Teléfono',
        'Correo', 'Estado'
    ]
    for col, encabezado in enumerate(encabezados, start=1):
        cell = ws.cell(row=2, column=col, value=encabezado)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 20

    # Datos
    alt_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    for fila, paciente in enumerate(pacientes, start=3):
        datos = [
            paciente.dni,
            paciente.nombre,
            paciente.apellido,
            paciente.fecha_nacimiento.strftime('%d/%m/%Y'),
            paciente.calcular_edad(),
            paciente.get_genero_display(),
            paciente.tipo_sangre or '-',
            paciente.telefono,
            paciente.email,
            'Activo' if paciente.activo else 'Inactivo',
        ]
        for col, valor in enumerate(datos, start=1):
            cell = ws.cell(row=fila, column=col, value=valor)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if fila % 2 == 0:
                cell.fill = alt_fill

    # Ancho de columnas
    anchos = [15, 18, 18, 18, 8, 12, 12, 15, 28, 10]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─── PDF ─────────────────────────────────────────────────────────────────────

def exportar_historia_pdf(paciente, historias):
    """
    Genera un PDF con la historia clínica de un paciente.
    Retorna bytes listos para enviar como respuesta HTTP.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer,
            Table, TableStyle, HRFlowable
        )
    except ImportError:
        raise ImportError("Instala reportlab: pip install reportlab")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    story = []

    # Título
    titulo_style = ParagraphStyle(
        'titulo',
        parent=styles['Title'],
        fontSize=18,
        textColor=colors.HexColor('#1a6b9a'),
        spaceAfter=6,
    )
    story.append(Paragraph("Historia Clínica", titulo_style))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1a6b9a')))
    story.append(Spacer(1, 0.4 * cm))

    # Datos del paciente
    subtitulo_style = ParagraphStyle(
        'subtitulo',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#1a6b9a'),
    )
    story.append(Paragraph("Datos del Paciente", subtitulo_style))

    datos_paciente = [
        ['DNI / Cédula:', paciente.dni, 'Nombre completo:', paciente.nombre_completo()],
        ['Fecha de nacimiento:', paciente.fecha_nacimiento.strftime('%d/%m/%Y'),
         'Edad:', f'{paciente.calcular_edad()} años'],
        ['Género:', paciente.get_genero_display(),
         'Tipo de sangre:', paciente.tipo_sangre or '-'],
        ['Teléfono:', paciente.telefono, 'Correo:', paciente.email],
    ]

    tabla_paciente = Table(datos_paciente, colWidths=[4 * cm, 6 * cm, 4 * cm, 6 * cm])
    tabla_paciente.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#EBF5FB')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#EBF5FB')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(tabla_paciente)
    story.append(Spacer(1, 0.6 * cm))

    # Historial de consultas
    story.append(Paragraph("Historial de Consultas", subtitulo_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#cccccc')))
    story.append(Spacer(1, 0.3 * cm))

    if not historias:
        story.append(Paragraph("No hay registros clínicos disponibles.", styles['Normal']))
    else:
        for i, historia in enumerate(historias, start=1):
            # Encabezado de consulta
            consulta_header = [
                [f'Consulta #{i}', historia.fecha.strftime('%d/%m/%Y'),
                 'Médico:', historia.medico_nombre or 'No especificado']
            ]
            tabla_header = Table(consulta_header, colWidths=[4 * cm, 4 * cm, 3 * cm, 9 * cm])
            tabla_header.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1a6b9a')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('PADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(tabla_header)

            # Contenido de la consulta
            contenido = [
                ['Motivo de consulta:', historia.motivo_consulta],
                ['Diagnóstico:', historia.diagnostico],
                ['Tratamiento:', historia.tratamiento or '-'],
                ['Observaciones:', historia.observaciones or '-'],
            ]
            tabla_contenido = Table(contenido, colWidths=[4 * cm, 16 * cm])
            tabla_contenido.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
                ('PADDING', (0, 0), (-1, -1), 5),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(tabla_contenido)
            story.append(Spacer(1, 0.4 * cm))

    # Pie de página
    story.append(Spacer(1, 0.5 * cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    pie_style = ParagraphStyle('pie', parent=styles['Normal'], fontSize=8, textColor=colors.grey)
    story.append(Paragraph(
        f"Documento generado el {date.today().strftime('%d/%m/%Y')} - Sistema de Gestión Hospitalaria",
        pie_style
    ))

    doc.build(story)
    buffer.seek(0)
    return buffer.read()
